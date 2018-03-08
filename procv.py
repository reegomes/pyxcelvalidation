from openpyxl import load_workbook
wb = load_workbook(filename='test.xlsx', read_only=False)
ws = wb['Sheet']

# Contador de linha e delimitador de linhas da planilha
line = 1
lines = 10

# Talvez essas linhas aqui estão sendo inuteis, só talvez
check1 = ws.iter_cols(min_row=line, min_col=2, max_col=2, max_row=line)
check2 = ws.iter_cols(min_row=line, min_col=8, max_col=8, max_row=line)
nnfe = ws.iter_cols(min_row=line, min_col=7, max_col=7, max_row=line)

# Listas
nomes1 = []
nomes2 = []
linhanfe = []


for col in ws.iter_cols(min_row=1, min_col=8, max_col=8, max_row=lines):
    for cell in col:
        line += 1
        check1 = cell.value
        
for col in ws.iter_cols(min_row=1, min_col=2, max_col=2, max_row=lines):
    for cell in col:
        check2 = cell.value
        
for col in ws.iter_cols(min_row=1, min_col=7, max_col=7, max_row=lines):
    for cell in col:
        nfe = ws.iter_cols(min_row=line, min_col=7, max_col=7, max_row=line)
        nnfe = (cell.value)

if (check1 == check2):
    print('Valores iguais para', check1,' numero da nota', nnfe,'.')
else:
    print("Diferentes")

#for n in nomes:
#    print(n)
